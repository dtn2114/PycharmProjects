__author__ = 'yutingchen'

from bs4 import BeautifulSoup
from global_setting_parsing import *
import csv
from openpyxl import Workbook, load_workbook


def initialize_scheduleD():
    try:
        wb = Workbook()
        ws = wb.create_sheet()
        ws.title = 'ScheduleD_BasicInfo'
        ws.append(basic_col_header)

        ws2 = wb.create_sheet()
        ws2.title = 'ScheduleD_book_records'
        ws2.append(ScheduleD_1L_col_header)

        ws3 = wb.create_sheet()
        ws3.title = 'ScheduleD_foreign_regulatory'
        ws3.append(ScheduleD_1M_col_header)

        ws4 = wb.create_sheet()
        ws4.title = 'ScheduleD_Section2'
        ws4.append(ScheduleD_2_col_header)

        ws5 = wb.create_sheet()
        ws5.title = 'ScheduleD_Advisers'
        ws5.append(ScheduleD_5G3_col_header)

        ws_successions = wb.create_sheet()
        ws_successions.title = 'ScheduleD_Section4'
        ws_successions.append(ScheduleD_4_col_header)

        ws6 = wb.create_sheet()
        ws6.title = 'ScheduleD_Section6B'
        ws6.append(ScheduleD_6B_col_header)

        ws7 = wb.create_sheet()
        ws7.title = 'ScheduleD_7A'
        ws7.append(ScheduleD_7A_col_header)

        ws8 = wb.create_sheet()
        ws8.title = 'ScheduleD_7B'
        ws8.append(ScheduleD_7B1_col_header)

        ws9 = wb.create_sheet()
        ws9.title = 'ScheduleD_Section7B(2)'
        ws9.append(ScheduleD_7B2_col_header)

        ws10 = wb.create_sheet()
        ws10.title = 'ScheduleD_Section9C'
        ws10.append(ScheduleD_9C_col_header)

        ws11 = wb.create_sheet()
        ws11.title = 'ScheduleD_Section10A'
        ws11.append(ScheduleD_10A_col_header)


        ws12 = wb.create_sheet()
        ws12.title = 'ScheduleD_Section10B'
        ws12.append(ScheduleD_10B_col_header)

        ws13 = wb.create_sheet()
        ws13.title = 'ScheduleD_Miscellaneous'
        ws13.append(ScheduleD_10Mis_col_header)

        wb.save(filename='output/scheduleD.xls')

        '''
        wb_re_read = load_workbook(filename='example.xls')
        sheet = wb_re_read.get_sheet_by_name('ScheduleD_BasicInfo')
        sheet.append(data)
        wb_re_read.save(filename='example2.xls')'''
    except:
        raise

def write_to_sheet(data,sheetname):
    try:
        wb_re_read = load_workbook(filename='output/scheduleD.xls')
        sheet = wb_re_read.get_sheet_by_name(sheetname)
        sheet.append(data)
        wb_re_read.save(filename='output/scheduleD.xls')
    except:
        print "Write to excel file suspended due to keyboard interruption"

def write_scheduleD_xlsx(file,comp_name):
    try:
        soup = BeautifulSoup(open(file))
        data = extract_ScheduleD_basic(comp_name,soup)
        write_to_sheet(data,'ScheduleD_BasicInfo')


        data = extract_scheduleD_1L(comp_name,soup)
        for entry in data:
            write_to_sheet(entry,'ScheduleD_book_records')

        data = extract_scheduleD_1M(comp_name,soup)
        for entry in data:
            write_to_sheet(entry,"ScheduleD_foreign_regulatory")

        data = extract_scheduleD_2(comp_name,soup)
        write_to_sheet(data,"ScheduleD_Section2")

        data = extract_scheduleD_4(comp_name,soup)
        write_to_sheet(data,"ScheduleD_Section4")

        data = extract_scheduleD_5G3(comp_name,soup)
        for entry in data:
            write_to_sheet(entry,"ScheduleD_Advisers")

        data = extract_scheduleD_6B(comp_name,soup)
        write_to_sheet(data,"ScheduleD_Section6B")

        data = extract_scheduleD_7A(comp_name,soup)
        for entry in data:
            write_to_sheet(entry,"ScheduleD_7A")

        data = extract_scheduleD_7B1(comp_name,soup)
        for entry in data:
            write_to_sheet(entry,"ScheduleD_7B")

        data = extract_scheduleD_7B2(comp_name,soup)
        for entry in data:
            write_to_sheet(entry,"ScheduleD_Section7B(2)")

        data = extract_scheduleD_9C(comp_name,soup)
        for entry in data:
            write_to_sheet(entry,"ScheduleD_Section9C")

        data = extract_scheduleD_10A(comp_name,soup)
        write_to_sheet(data,"ScheduleD_Section10A")

        data = extract_scheduleD_10B(comp_name,soup)
        write_to_sheet(data,"ScheduleD_Section10B")

        data = extract_scheduleD_10mis(comp_name,soup)
        write_to_sheet(data,"ScheduleD_Miscellaneous")

    except:
        raise

'''
def write_scheduleD_xlsx(filename,file):
    """create csv file for schedule A's """
    try:
        wb = xlwt.Workbook()
        ws = wb.add_sheet('ScheduleD_BasicInfo')
        for i in range(0,len(basic_col_header)):
            ws.write(0,i,basic_col_header[i])
        soup = BeautifulSoup(open(file))
        data = extract_ScheduleD_basic(soup)

        for i in range(0,len(data)):
            ws.write(1,i,data[i])

        ws2 = wb.add_sheet("ScheduleD_book_records")
        data = extract_scheduleD_1L("Morgan Stanley",soup)
        for i in range(0,len(ScheduleD_1L_col_header)):
            ws2.write(0,i,ScheduleD_1L_col_header[i])

        for j in range(0,len(data)):
            for k in range(0,len(data[j])):
                ws2.write(j+1,k,data[j][k])

        ws3 = wb.add_sheet("ScheduleD_foreign_regulatory")
        data = extract_scheduleD_1M("Morgan Stanley",soup)
        for i in range(0,len(ScheduleD_1M_col_header)):
            ws3.write(0,i,ScheduleD_1M_col_header[i])

        for j in range(0,len(data)):
            for k in range(0,len(data[j])):
                ws3.write(j+1,k,data[j][k])

        ws4 = wb.add_sheet("ScheduleD_Section2")
        data = extract_scheduleD_2(data[0],soup)
        for i in range(0,len(ScheduleD_2_col_header)):
            ws4.write(0,i,ScheduleD_2_col_header[i])

        for k in range(0,len(data)):
            ws4.write(1,k,data[k])

        ws5 = wb.add_sheet("ScheduleD_Advisers")
        data = extract_scheduleD_5G3("Morgan Stanley",soup)
        for i in range(0,len(ScheduleD_5G3_col_header)):
            ws5.write(0,i,ScheduleD_5G3_col_header[i])

        for j in range(0,len(data)):
            for k in range(0,len(data[j])):
                ws5.write(j+1,k,data[j][k])

        ws6 = wb.add_sheet("ScheduleD_Section6B")
        data = extract_scheduleD_6B("Morgan Stanley",soup)
        for i in range(0,len(ScheduleD_6B_col_header)):
            ws6.write(0,i,ScheduleD_6B_col_header[i])

        for k in range(0,len(data)):
            ws6.write(1,k,data[k])

        ws7 = wb.add_sheet("ScheduleD_7A")
        data = extract_scheduleD_7A("Morgan Stanley",soup)
        for i in range(0,len(ScheduleD_7A_col_header)):
            ws7.write(0,i,ScheduleD_7A_col_header[i])

        for j in range(0,len(data)):
            for k in range(0,len(data[j])):
                ws7.write(j+1,k,data[j][k])

        ws7 = wb.add_sheet("ScheduleD_7A")
        extract_scheduleD_7B1("Morgan Stanley",soup)

        ws8 = wb.add_sheet("ScheduleD_7B")
        data = extract_scheduleD_7B1("Morgan Stanley",soup)
        for i in range(0,len(ScheduleD_7B_col_header)):
            ws8.write(0,i,ScheduleD_7B_col_header[i])

        for j in range(0,len(data)):
            for k in range(0,len(data[j])):
                ws8.write(j+1,k,data[j][k])


        ws9 = wb.add_sheet("ScheduleD_Section7B(2)")
        data = extract_scheduleD_7B2("Morgan Stanley",soup)
        for i in range(0,len(ScheduleD_7B2_col_header)):
            ws9.write(0,i,ScheduleD_7B2_col_header[i])

        for j in range(0,len(data)):
            for k in range(0,len(data[j])):
                ws9.write(j+1,k,data[j][k])

        ws10 = wb.add_sheet("ScheduleD_Section9C")
        data = extract_scheduleD_9C("Morgan Stanley",soup)
        for i in range(0,len(ScheduleD_9C_col_header)):
            ws10.write(0,i,ScheduleD_9C_col_header[i])

        for j in range(0,len(data)):
            for k in range(0,len(data[j])):
                ws10.write(j+1,k,data[j][k])


        #skip 10A
        ws12 = wb.add_sheet("ScheduleD_Section10B")
        data = extract_scheduleD_10B("Morgan Stanley",soup)
        for i in range(0,len(ScheduleD_10B_col_header)):
            ws12.write(0,i,ScheduleD_10B_col_header[i])
        for k in range(0,len(data)):
            ws12.write(1,k,data[k])


        ws13 = wb.add_sheet("ScheduleD_Miscellaneous")
        data = extract_scheduleD_10mis("Morgan Stanley",soup)
        for i in range(0,len(ScheduleD_10Mis_col_header)):
            ws13.write(0,i,ScheduleD_10Mis_col_header[i])

        for k in range(0,len(data)):
            ws13.write(1,k,data[k])


        wb.save(filename)

        writer = csv.writer(outfile, delimiter=',', quoting=csv.QUOTE_ALL)
        for row in col_header:
            writer.writerow(row)
        outfile.close()

    except:
        raise '''


def extract_ScheduleD_basic(comp_name,soup):
    data = []
    #1B
    scheduleD_1Bnone = soup.find('span',{'id':SCHEDULE_D_NO_NAMES})
    if scheduleD_1Bnone!=None:
        data.append(comp_name)
        data.append("No other names")
        for i in range(0,55):
            data.append("N/A")
    else:
        scheduleD_1b = soup.find('span',{'id':SCHEDULE_D_NAMES})
        scheduleD_1b_table = scheduleD_1b.parent.parent.parent.find_next('tr')
        D1b_table = scheduleD_1b_table.find('table',{'class':'PaperFormTableData'})
        if D1b_table!=None:
            data.append(comp_name)
            name_span = D1b_table.find("span")
            data.append(name_span.get_text())
            imgs = D1b_table.find_all("img")
            for img in imgs:
                if "not" not in img['alt']:
                    data.append(1)
                else:
                    data.append(0)
            other_span = imgs[len(imgs)-1].find_next("span")
            data.append(other_span.get_text())

    #1F
    scheduleD_1Fnone = soup.find('span',{'id':SCHEDULE_D_NO_OFFICES})
    if scheduleD_1Fnone!=None:
        data.append("No other offices")
        for i in range(0,8):
            data.append("N/A")
    else:
        scheduleD_1f = soup.find('span',{'id':SCHEDULE_D_OFFICES})
        scheduleD_1f_table = scheduleD_1f.parent.parent.parent.find_next('tr')
        D1f_table = scheduleD_1f_table.find('table',{'class':'PaperFormTableData'})
        spans = D1f_table.find_all("span")
        for i in range(0,len(spans)):
            if i !=3 and i!=5: #remove an empty span
                data.append(spans[i].get_text())
        private_address = D1f_table.find("img")
        if "not" not in private_address['alt']:
            data.append(1)
        else:
            data.append(0)

    #1I
    scheduleD_1I = soup.find('tr',{'id':SCHEDULE_D_WEBSITE})
    if scheduleD_1I!=None:
        D1I_span = scheduleD_1I.find_next('tr').find_next('tr').find("span")
        data.append(D1I_span.get_text())
    else:
        data.append("None")

    return data

def extract_scheduleD_1L(comp_name,soup):
     #1L
    data = []
    scheduleD_1L = soup.find('span',{'id':SCHEDULE_D_RECORDS})
    scheduleD_1L_table = scheduleD_1L.parent.parent.parent.find_next('tr')
    D1L_tables = scheduleD_1L_table.find_all('table',{'class':'PaperFormTableData'})
    for D1L_table in D1L_tables:
        rows = []
        rows.append(comp_name)
        spans = D1L_table.find_all("span")
        for span in spans:
            rows.append(span.get_text())
        imgs = D1L_table.find_all("img")

        if "not" not in imgs[0]['alt']:
            rows.append("not private")
        else:
            rows.append("private")
        if "changed" in imgs[1]['alt']:
            rows.append("one of your branch offices or affiliates")
        elif "changed" in imgs[2]['alt']:
            rows.append("a third-party unaffiliated recordkeeper")
        elif "changed" in imgs[3]['alt']:
            rows.append("other")
        data.append(rows)
    print data
    return data

def extract_scheduleD_1M(comp_name,soup):
    data = []
    #1M
    scheduleD_1Mnone = soup.find('span',{'id':SCHEDULE_D_NO_FOREIGN})
    if scheduleD_1Mnone!=None:
        row = []
        row.append(comp_name)
        row.append("No Info Registration with Foreign Authorities")
        data.append(row)
    else:
        scheduleD_1m = soup.find('span',{'id':SCHEDULE_D_FOREIGN_REG})
        scheduleD_1m_table = scheduleD_1m.parent.parent.parent.find_next('tr')
        D1m_tables = scheduleD_1m_table.find_all('table',{'class':'PaperFormTableData'})
        for D1m_table in D1m_tables:
            rows = []
            rows.append(comp_name)
            spans = D1m_table.find_all("span")
            for span in spans:
                rows.append(span.get_text())
            data.append(rows)
    return data


def extract_scheduleD_2(comp_name,soup):
    data = []
    data.append(comp_name)
    #2A8
    scheduleD_2A8 = soup.find('tr',{'id':SCHEDULE_D_2A8})

    if scheduleD_2A8 != None:
        scheduleD_2A8_table = scheduleD_2A8.find_next("tr").find('table',{'class':'PaperFormTableData'})
        spans = scheduleD_2A8_table.find_all("span")
        for span in spans:
            data.append(span.get_text().encode('ascii', 'ignore'))
        #2A9
        scheduleD_2A9 = soup.find('tr',{'id':SCHEDULE_D_2A9})
        scheduleD_2A9_table = scheduleD_2A9.find_next("tr").find('table',{'class':'PaperFormTableData'})
        imgs = scheduleD_2A9_table.find_all("img")
        for img in imgs:
            if "not" not in img['alt']:
                data.append(1)
            else:
                data.append(0)

        #2A10
        scheduleD_2A10 = soup.find('tr',{'id':SCHEDULE_D_2A10})
        scheduleD_2A10_table = scheduleD_2A10.find_next("tr").find('table',{'class':'PaperFormTableData'})
        imgs = scheduleD_2A10_table.find_all("img")
        for img in imgs:
            if "not" not in img['alt']:
                data.append(1)
            else:
                data.append(0)

        #2A12
        scheduleD_2A12 = soup.find('tr',{'id':SCHEDULE_D_2A12})
        scheduleD_2A12_table = scheduleD_2A12.find_next("tr").find('table',{'class':'PaperFormTableData'})
        spans = scheduleD_2A12_table.find_all("span")
        for span in spans:
            data.append(span.get_text().encode('ascii', 'ignore'))
    else:
        data.append("No Info")
    return data

def extract_scheduleD_4(comp_name,soup):
    data = []
    data.append(comp_name)
    #4 Succession
    scheduleD_4None = soup.find('span',{'id':SCHEDULE_D_NO_SUCCESSION})
    if scheduleD_4None!=None:
        data.append('No Succession')
    else:
        data.append("Have succession")
    return data

def extract_scheduleD_5G3(comp_name,soup):
    data = []
    #5G
    scheduleD_5G = soup.find('span',{'id':SCHEDULE_D_NO_5G})
    if scheduleD_5G!=None:
        row = []
        row.append(comp_name)
        row.append('No Advisers')
        data.append(row)
    else:

        scheduleD_5G = soup.find('span',{'id':SCHEDULE_D_5G})
        tables = scheduleD_5G.parent.parent.parent.parent.find_all('table',{'class':'PaperFormTableData'})
        for table in tables:
            row = []
            row.append(comp_name)
            spans = table.find_all("span")
            sec = ''
            for span in spans:
                sec = sec+span.get_text()
            row.append(sec)
            data.append(row)

    return data

def extract_scheduleD_6B(comp_name,soup):
    data = []
    #6B(2)
    scheduleD_6B2 = soup.find('tr',{'id':SCHEDULE_D_6B2})
    table = scheduleD_6B2.parent.parent.find('table',{'class':'PaperFormTableData'})
    data.append(comp_name)
    spans = table.find_all("span")

    for span in spans:
        data.append(span.get_text())

    scheduleD_6B3 = soup.find('tr',{'id':SCHEDULE_D_6B3})
    table = scheduleD_6B3.parent.parent.find('table',{'class':'PaperFormTableData'})
    spans = table.find_all("span")

    for span in spans:
        data.append(span.get_text())

    return data

def extract_scheduleD_7A(comp_name,soup):
    data = []
    #7A
    scheduleD_7Anone = soup.find('span',{'id':SCHEDULE_D_NO_7A})
    if scheduleD_7Anone!=None:
        row = []
        row.append(comp_name)
        row.append("No  Financial Industry Affiliations")
        data.append(row)
    else:
        scheduleD_7A = soup.find('span',{'id':SCHEDULE_D_7A})
        tables = scheduleD_7A.parent.parent.parent.parent.find_all('table',{'class':'flatBorderTable'})
        for table in tables:
            row = []
            row.append(comp_name)
            spans = table.find_all("span")

            row.append(spans[0].get_text().encode('ascii', 'ignore'))
            row.append(spans[1].get_text().encode('ascii', 'ignore'))

            sec = spans[3].get_text().encode('ascii', 'ignore')+'-'+spans[4].get_text().encode('ascii', 'ignore')
            row.append(sec)

            imgs = table.find_all('img')
            for i in range(0,16):
                if "not" not in imgs[i]['alt']:
                    row.append(1)
                else:
                    row.append(0)

            for j in [16,18,20,22,25,27,29,31]:
                if "not" not in imgs[j]['alt']:
                    row.append(1)
                else:
                    row.append(0)

            #fit the names from the regulatory list into one cell
            regu_list = table.find('tr',{'class':'QueryTableTitle'})
            if regu_list != None:
                table_results  = regu_list.parent.find_all('tr',{'class':'PrintHistRed'})
                list = []
                for result in table_results:
                    list.append(result.get_text().encode('ascii', 'ignore'))
                row.append(list)
            data.append(row)
    return data

def extract_scheduleD_7B1(comp_name,soup):
    data = []
    #7B1
    scheduleD_7Bnone = soup.find('span',{'id':SCHEDULE_D_NO_7B})
    if scheduleD_7Bnone!=None:
        row = []
        row.append(comp_name)
        row.append("No Private Fund Reporting")
        data.append(row)
    else:
        scheduleD_7B = soup.find('span',{'id':SCHEDULE_D_7B})
        tables = scheduleD_7B.parent.parent.parent.parent.find_all('table',{'class':'PaperFormTableData'})
        for table in tables:
            row = []
            row.append(comp_name)
            imgs = table.find_all('img')
            #make sure it's not subtables within the section
            if len(imgs)>40:
                spans = table.find_all("span")
                for i in range(0,4):
                    row.append(spans[i].get_text().encode('ascii', 'ignore'))

                list_of_names_Q3 = []
                list_of_names_Q5 = []
                list_of_names_Q6 = []
                list_of_names_Q7 = []

                list_tables = table.find_all("table")
                for list_table in list_tables:
                    if list_table.has_attr('id'):
                        if SCHEDULE_D_7B_Q3 in list_table['id'] and SCHEDULE_D_7B_TABLE in list_table['id']:
                            trs = list_table.find_all('tr',{'class':'PrintHistRed'})
                            for tr in trs:
                                td_text = tr.find("td").get_text().encode('ascii', 'ignore')
                                list_of_names_Q3.append(td_text)

                        elif SCHEDULE_D_7B_Q5 in list_table['id'] and SCHEDULE_D_7B_TABLE in list_table['id']:
                            trs = list_table.find_all('tr',{'class':'PrintHistRed'})
                            for tr in trs:
                                td_text = tr.find("td").get_text().encode('ascii', 'ignore')
                                list_of_names_Q5.append(td_text)
                        elif SCHEDULE_D_7B_Q6 in list_table['id'] and SCHEDULE_D_7B_TABLE in list_table['id']:
                            trs = list_table.find_all('tr',{'class':'PrintHistRed'})
                            for tr in trs:
                                td_text = tr.find("td").get_text().encode('ascii', 'ignore')
                                list_of_names_Q6.append(td_text)
                        elif SCHEDULE_D_7B_Q7 in list_table['id'] and SCHEDULE_D_7B_TABLE in list_table['id']:
                            trs = list_table.find_all('tr',{'class':'PrintHistRed'})
                            for tr in trs:
                                td_text = tr.find("td").get_text().encode('ascii', 'ignore')
                                list_of_names_Q7.append(td_text)
                q3_names =""
                for names in list_of_names_Q3:
                    q3_names = q3_names + names+";"
                if q3_names == "":
                    q3_names = "No Information Filed"
                row.append(q3_names)

                #q4
                for i in range(0,2):
                    if 'not' not in imgs[i]['alt']:
                        row.append(1)
                    else:
                        row.append(0)

                q5_names =""
                for names in list_of_names_Q5:
                    q5_names = q5_names + names+";"
                if q5_names == "":
                    q5_names = "No Information Filed"
                row.append(q5_names)

                #q6
                for j in [2,4]:
                    if 'not' not in imgs[j]['alt']:
                        row.append(1)
                    else:
                        row.append(0)

                q6_names =""
                for names in list_of_names_Q6:
                    q6_names = q6_names + names+";"
                if q6_names == "":
                    q6_names = "No Information Filed"
                row.append(q6_names)

                #q7
                q7_names =""
                for names in list_of_names_Q7:
                    q7_names = q7_names + names+";"
                if q7_names == "":
                    q7_names = "No Information Filed"
                row.append(q7_names)

                #q8,q9
                for k in [6,8,10]:
                    if 'not' not in imgs[k]['alt']:
                        row.append(1)
                    else:
                        row.append(0)

                #q10
                for l in range(12,19):
                    if 'not' not in imgs[l]['alt']:
                        if l == 12:
                            row.append("hedge fund")
                        elif l ==13:
                            row.append("liquidity fund")
                        elif l==14:
                            row.append("private equity fund")
                        elif l==15:
                            row.append("real estate fund")
                        elif l==16:
                            row.append("securitized asset fund")
                        elif l==17:
                            row.append("venture capital fund")
                        elif l==18:
                            row.append("other")
                #q11
                q11_span = imgs[18].find_next("span").find_next("span")
                q11_span_text = q11_span.get_text().encode('ascii', 'ignore')
                row.append(q11_span_text)

                #q12-q16
                q_span = q11_span
                for i in range(0,5):
                    q_span = q_span.find_next("span")
                    q_span_text = q_span.get_text().encode('ascii', 'ignore')
                    row.append(q_span_text)

                #q17
                q17_img = imgs[19]
                if "not" not in q17_img['alt']:
                    row.append(1)
                else:
                    row.append(0)

                list_of_names_Q17 = []
                list_of_names_Q18 = []
                list_of_names_Q22 = []
                for list_table in list_tables:
                    if list_table.has_attr('id'):
                        if SCHEDULE_D_7B_Q17 in list_table['id'] and SCHEDULE_D_7B_TABLE in list_table['id']:
                            trs = list_table.find_all('tr',{'class':'PrintHistRed'})
                            for tr in trs:
                                td_text = tr.find("td").get_text().encode('ascii', 'ignore')
                                list_of_names_Q17.append(td_text)
                        elif SCHEDULE_D_7B_Q18 in list_table['id'] and SCHEDULE_D_7B_TABLE in list_table['id']:
                            trs = list_table.find_all('tr',{'class':'PrintHistRed'})
                            for tr in trs:
                                td_text = tr.find("td").get_text().encode('ascii', 'ignore')
                                list_of_names_Q18.append(td_text)
                        elif SCHEDULE_D_7B_Q22 in list_table['id'] and SCHEDULE_D_7B_TABLE in list_table['id']:
                            trs = list_table.find_all('tr',{'class':'PrintHistRed'})
                            for tr in trs:
                                td_text = tr.find("td").get_text().encode('ascii', 'ignore')
                                list_of_names_Q22.append(td_text)
                q17_names =""
                for names in list_of_names_Q17:
                    q17_names = q17_names + names+";"
                if q17_names == "":
                    q17_names = "No Information Filed"
                row.append(q17_names)

                #q18
                q18_img = imgs[21]
                if "not" not in q18_img['alt']:
                    row.append(1)
                else:
                    row.append(0)

                q18_names =""
                for names in list_of_names_Q18:
                    q18_names = q18_names + names+";"
                if q18_names == "":
                    q18_names = "No Information Filed"
                row.append(q18_names)

                #q19
                q19_img = imgs[23]
                if "not" not in q19_img['alt']:
                    row.append(1)
                else:
                    row.append(0)

                #q20
                q20_span = imgs[24].find_next("span")
                q20_span_text = q20_span.get_text().encode('ascii', 'ignore')
                row.append(q20_span_text)

                #q21
                q21_img = imgs[25]
                if "not" not in q21_img['alt']:
                    row.append(1)
                else:
                    row.append(0)
                #q22
                q22_names =""
                for names in list_of_names_Q22:
                    names = names.strip("\n")
                    q22_names = q22_names + names+";"
                if q22_names == "":
                    q22_names = "No Information Filed"
                row.append(q22_names)
                #row.append(list_of_names_Q22)

                #q23a,23b
                q23a_img = imgs[27]
                if "not" not in q23a_img['alt']:
                    row.append(1)
                else:
                    row.append(0)
                q23b_img = imgs[29]
                if "not" not in q23b_img['alt']:
                    row.append(1)
                else:
                    row.append(0)

                num_q23_table = 0 #used to keep track of #of tabless and therefore track the index of imgs
                if "not" not in q23a_img['alt']:
                    #find table info
                    q23_table = q23b_img.parent.find_next('tr').find_next('tr').find_next('tr').find('table',{'class':'flatBorderTable'})
                    span = q23_table.find("span").get_text().encode('ascii', 'ignore')
                    number = filter(str.isdigit, span)
                    num_q23_table = int(number)
                    row.append(span)
                else:
                    row.append("N/A")

                #q23g,h
                img_index_adjust = 6*num_q23_table
                q23g_img = imgs[31+img_index_adjust]
                q23g_img2 = imgs[32+img_index_adjust]
                if "not" not in q23g_img['alt']:
                    row.append(1)
                elif "not" not in q23g_img2['alt'] :
                    row.append(0)
                else:
                    row.append("N/A")

                if "not" not in imgs[33+img_index_adjust]['alt']:
                    row.append("Yes")
                elif "not" not in imgs[34+img_index_adjust]['alt']:
                    row.append("No")
                elif "not" not in imgs[35+img_index_adjust]['alt']:
                    row.append("Report Not Yet Received")
                else:
                    row.append("N/A")

                #24
                q24_img = imgs[36+img_index_adjust]
                num_q24_table = 0 #keep track of the img index for later sections
                if "not" not in q24_img['alt']:
                    row.append(1)
                    #find table info
                    q24_table = q24_img.parent.parent.find_next('tr').find_next('tr').find('table',{'class':'flatBorderTable'})
                    span = q24_table.find("span").get_text().encode('ascii', 'ignore')
                    number = filter(str.isdigit, span)
                    num_q24_table = int(number)
                    row.append(span)

                else:
                    row.append(0)
                    row.append("N/A")

                #25
                img_index_adjust2 = 2*num_q24_table
                q25_img = imgs[38+img_index_adjust+img_index_adjust2]

                num_q25_table = 0
                if "not" not in q25_img['alt']:
                    row.append(1)
                    #find table info
                    q25_table = q25_img.parent.parent.find_next('tr').find_next('tr').find('table',{'class':'flatBorderTable'})
                    span = q25_table.find("span").get_text().encode('ascii', 'ignore')
                    number = filter(str.isdigit, span)
                    num_q25_table = int(number)
                    row.append(span)

                else:
                    row.append(0)
                    row.append("N/A")

                #26
                img_index_adjust3 = 2*num_q25_table
                q26_img = imgs[40+img_index_adjust+img_index_adjust2+img_index_adjust3]
                num_q26_table = 0
                if "not" not in q26_img['alt']:
                    row.append(1)
                    #find table info
                    q26_table = q26_img.parent.parent.find_next('tr').find_next('tr').find('table',{'class':'flatBorderTable'})
                    span = q26_table.find("span").get_text().encode('ascii', 'ignore')
                    number = filter(str.isdigit, span)
                    num_q26_table = int(number)
                    row.append(span)

                else:
                    row.append(0)
                    row.append("N/A")

                #27,28
                img_index_adjust4 = 5*num_q26_table
                q28_img = imgs[42+img_index_adjust+img_index_adjust2+img_index_adjust3+img_index_adjust4]
                q27_tr = q28_img.parent.parent.find_previous('tr').find_previous('tr').find_previous('tr').find_previous('tr').find_previous('tr')
                q27_span = q27_tr.find('span')
                if q27_span!= None:
                    q27_span_text = q27_span.get_text().encode('ascii', 'ignore')
                else:
                    q27_span_text = 'N/A'
                row.append(q27_span_text)

                if "not" not in q28_img['alt']:
                    row.append(1)
                    #find table info
                    q28_table = q28_img.parent.parent.find_next('tr').find_next('tr').find('table',{'class':'flatBorderTable'})
                    span = q28_table.find("span").get_text().encode('ascii', 'ignore')
                    number = filter(str.isdigit, span)
                    row.append(span)
                else:
                    row.append(0)
                    row.append("N/A")

                data.append(row)
        print data
    return data

def extract_scheduleD_7B2(comp_name,soup):
    data = []
    #7B2
    scheduleD_7B2none = soup.find('span',{'id':SCHEDULE_D_NO_7B2})
    if scheduleD_7B2none!=None:
        row = []
        row.append(comp_name)
        row.append("No Private Fund Reporting(2)")
        data.append(row)

    else:
        scheduleD_7B2 = soup.find('span',{'id':SCHEDULE_D_7B2})
        tables = scheduleD_7B2.parent.parent.parent.parent.find_all('table',{'class':'PaperFormTableData'})
        for table in tables:
            row = []
            row.append(comp_name)
            spans = table.find_all("span")
            for span in spans:
                row.append(span.get_text().encode('ascii', 'ignore'))
            img = table.find("img")
            if "not" not in img['alt']:
                row.append(1)
            else:
                row.append(0)
            data.append(row)

    print data
    return data


def extract_scheduleD_9C(comp_name,soup):
    data = []
    #9C
    scheduleD_9Cnone = soup.find('span',{'id':SCHEDULE_D_NO_9C})
    if scheduleD_9Cnone!=None:
        row = []
        row.append(comp_name)
        row.append("No Independent Public Accountant")
        data.append(row)
    else:
        scheduleD_9C = soup.find('span',{'id':SCHEDULE_D_9C})
        scheduleD_9C_tables = scheduleD_9C.parent.parent.parent.parent.find_all('table',{'class':'flatBorderTable'})
        for scheduleD_9C_table in scheduleD_9C_tables:
            row = []
            row.append(comp_name)
            spans = scheduleD_9C_table.find_all("span")
            for i in [0,1,2,3,5,7,8]:
                row.append(spans[i].get_text().encode('ascii', 'ignore'))


            imgs = scheduleD_9C_table.find_all("img")
            for j in [0,2]:
                if "not" not in imgs[j]['alt']:
                    row.append(1)
                else:
                    row.append(0)
            for k in range(4,7):
                if "not" not in imgs[k]['alt']:
                    row.append(1)
                else:
                    row.append(0)

            if "not" not in imgs[7]['alt']:
                row.append("yes")
            elif "not" not in imgs[8]['alt']:
                row.append("no")
            elif "not" not in imgs[9]['alt']:
                row.append("Report Not Yet Received")
            else:
                row.append("N/A")
            data.append(row)
        print data
    return data

def extract_scheduleD_10A(comp_name,soup):
    data = []
    data.append(comp_name)
    scheduleD_10Anone = soup.find('span',{'id':SCHEDULE_D_NO_10A})
    if scheduleD_10Anone!=None:
        data.append("No Control Person")
    else:
        data.append("Has Control Person")
    return data

def extract_scheduleD_10B(comp_name,soup):
    data = []
    #10B
    scheduleD_10Bnone = soup.find('span',{'id':SCHEDULE_D_NO_10B})
    if scheduleD_10Bnone!=None:
        data.append(comp_name)
        data.append("No Control Person Public Reporting Companies")

    else:
        scheduleD_10B = soup.find('span',{'id':SCHEDULE_D_10B})
        scheduleD_10B_table = scheduleD_10B.parent.parent.parent.parent.find('table',{'class':'flatBorderTable'})
        data.append(comp_name)
        spans = scheduleD_10B_table.find_all("span")
        for span in spans:
            data.append(span.get_text().encode('ascii', 'ignore'))
        print data

    return data

def extract_scheduleD_10mis(comp_name,soup):
    data = []
    #MISCELLANEOUS
    scheduleD_mis = soup.find('tr',{'id':SCHEDULE_D_MIS})
    mis_table = scheduleD_mis.parent.parent.find('table',{'class':'PaperFormTableData'})
    span = mis_table.find("span")
    data.append(comp_name)
    data.append(span.get_text().encode('ascii', 'ignore'))

    print data
    #add a label to those files with section 4
    return data